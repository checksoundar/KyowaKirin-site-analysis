#!/usr/bin/env python3
"""Build an Excel workbook mapping every analysed sumitclub.jp URL to its
category, migration mode and DOM template family.

Sheet 1 (URL Mapping): one row per URL.
Sheet 2 (Category Summary): counts per category.
Sheet 3 (Mode Summary): counts per migration mode.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PER_URL = '/workspace/current/.migration/sumitJP/per-url.json'
OUT = '/workspace/current/SumitclubJP_URL_Category_Mapping.xlsx'

rows = json.load(open(PER_URL))

# Family mapping (same grouping as the Word report).
FAMILY = {
    'Standard AEM Content Shell (/ja/)': [
        'Usage / How-to Page', 'Travel Service Page', 'Insurance Product/Info',
        'Point Program Page', 'Legal / Policy Text', 'Contact / Support',
        'Campaign Landing Page', 'Entertainment/Lifestyle Page', 'Gourmet Service Page',
        'Club Online Promo (unique)', 'Section Landing Page', 'Section Index',
        'Commercial Card Page', 'Other Content Page', 'Notice / News Detail',
        'Notice / News Index', 'Sitemap',
    ],
    'AEM Card-Detail / Listing Shell': [
        'Card Product Detail', 'Card Lineup Listing/Index',
    ],
    'AEM Corporate Shell (/ja/corporate_site)': [
        'Corporate Info (AEM)', 'Corporate News List', 'Corporate Special/Anniversary',
    ],
    'Legacy Corporate Microsite (/corporate)': [
        'Corporate Info (legacy shell)', 'Corporate Kaizen Notice (archive)',
        'Corporate Recruit Page',
    ],
    'Legacy English Shell (/en)': [
        'Homepage (locale)', 'Service Page (legacy en)', 'Announcement Page',
        'FAQ Page', 'Info Index', 'HTTP Error Page',
    ],
    'Standalone Microsites (non-AEM)': [
        'Card Application Landing Page (LP)', 'Identity Verification Microsite',
        'Club Online Bumper Page', 'Cancellation/Termination Flow',
    ],
    'Auth-gated / System (excluded)': [
        'Member Dining-Selection Detail (R####)', 'Member Dining-Selection Index/Search',
        'Login/Member System Page', 'Sign-on / Login System', 'System Error Stub',
    ],
    'Non-Page Assets (excluded)': [
        'SSI Include Fragment', 'LP Partial Fragment', 'Bumper/Redirect Stub',
        'Blank/Helper Stub', 'Tracking/System Stub', 'Modal/Popup Fragment',
        'Auth Widget Fragment', 'Redirect / Router Stub', 'App-bridge Stub',
    ],
    'E-statement Email Templates (separate track)': [
        'E-statement Email Template',
    ],
}
cat_to_family = {c: f for f, cs in FAMILY.items() for c in cs}

# Main-content DOM pattern per category. This describes the markup BETWEEN the
# header and footer Experience Fragments (i.e. the page's main content container
# and its repeated block sequence), derived from live DOM inspection.
# Standard AEM pages wrap main content in <div class="root responsivegrid">.
AEM_ROOT = 'div.root.responsivegrid'
DOM_PATTERN = {
    'Homepage':
        f'{AEM_ROOT} > hero/banner + recommended-card grid + service tiles + notice list '
        '(no breadcrumb). Blocks: hero, cardGrid, news list.',
    'Card Product Detail':
        f'{AEM_ROOT} > breadcrumb + h1 + image carousel + tab nav + accordion + spec table '
        '+ related-card carousel. Richest template (carousel+tabs+accordion+table).',
    'Card Lineup Listing/Index':
        f'{AEM_ROOT} > breadcrumb + h1 + 2-3 comparison tables + card carousel + card grid.',
    'Card Lineup Listing / Compare':
        f'{AEM_ROOT} > breadcrumb + h1 + comparison tables + card carousel.',
    'Section Landing Page':
        f'{AEM_ROOT} > breadcrumb + h1 + multiple h2 category blocks + card grid.',
    'Section Index':
        f'{AEM_ROOT} > breadcrumb + h1 + section card grid / link lists.',
    'Notice / News Detail':
        f'{AEM_ROOT} > breadcrumb + h1 + intro paragraph + optional table + notes list (rich text).',
    'Notice / News Index':
        f'{AEM_ROOT} > breadcrumb + h1 + category tab filter + year-grouped link lists (accordion).',
    'Usage / How-to Page':
        f'{AEM_ROOT} > breadcrumb + left category sub-nav + h1 + h2 feature sections (image+text, steps).',
    'Travel Service Page':
        f'{AEM_ROOT} > breadcrumb + left sub-nav + h1 + service card list + accordion (icon legend).',
    'Insurance Product/Info':
        f'{AEM_ROOT} > breadcrumb + left sub-nav + h1 + in-page tab anchors + comparison tables.',
    'Point Program Page':
        f'{AEM_ROOT} > breadcrumb + left sub-nav + h1 + h2/h3 content sections (rich text + cards).',
    'Campaign Landing Page':
        f'{AEM_ROOT} > breadcrumb + h1 + promo card blocks + CTA.',
    'Legal / Policy Text':
        f'{AEM_ROOT} > breadcrumb + h1 + numbered h3 headings + paragraphs (text-heavy, no blocks).',
    'Contact / Support':
        f'{AEM_ROOT} > breadcrumb + h1 + FAQ links + accordion of phone-contact sections (no real form).',
    'Entertainment/Lifestyle Page':
        f'{AEM_ROOT} > breadcrumb + left sub-nav + h1 + service/feature card sections.',
    'Gourmet Service Page':
        f'{AEM_ROOT} > breadcrumb + left sub-nav + h1 + service card sections.',
    'Club Online Promo (unique)':
        f'{AEM_ROOT} > h1 + promo image+text blocks + CTA (Club Online context).',
    'Commercial Card Page':
        f'{AEM_ROOT} > breadcrumb + h1 + procedure/step sections + tables.',
    'Other Content Page':
        f'{AEM_ROOT} > h1 + mixed rich-text / card sections (standard AEM shell).',
    'Sitemap':
        f'{AEM_ROOT} > h1 + ~11 category h2 + nested link lists.',
    'Corporate Info (AEM)':
        f'{AEM_ROOT} (corporate header/footer XF) > breadcrumb + h1 + content sections + tables.',
    'Corporate News List':
        f'{AEM_ROOT} (corporate chrome) > h1 + dated news link list.',
    'Corporate Special/Anniversary':
        f'{AEM_ROOT} (corporate chrome) > h1 + feature/story sections.',
    'Corporate Info (legacy shell)':
        'Legacy /corporate microsite: div.topPage / bespoke divs (NOT the AEM responsivegrid). '
        'Often a JS redirect stub to /ja/corporate_site.',
    'Corporate Kaizen Notice (archive)':
        'Legacy static page: single wrapper div + breadcrumb + h1 + dated kaizen notice body '
        '(reuses corporate chrome; no AEM responsivegrid).',
    'Corporate Recruit Page':
        'Recruit microsite (jQuery): section.main > hero carousel + content sections '
        '(separate codebase, not AEM shell).',
    'Card Application Landing Page (LP)':
        'Standalone marketing LP microsite: bespoke divs, multiple offer/contents sections, '
        'tables, iframes (assembled from _offer/_header/_footer partials). Not AEM shell.',
    'Identity Verification Microsite':
        'Static no-js microsite: div.wrapper + form scaffolding (built from honninkakunin '
        '/include/* partials). Not AEM shell.',
    'Club Online Bumper Page':
        'Interstitial bumper: minimal wrapper div + message + continue CTA.',
    'Cancellation/Termination Flow':
        'Minimal-chrome flow page: wrapper div + step content + CTA.',
    'Homepage (locale)':
        'Legacy /en shell: div.site-width content columns + carousel (xhtml, NOT AEM responsivegrid).',
    'Service Page (legacy en)':
        'Legacy /en shell: div.site-width content + side nav.',
    'Announcement Page':
        'Legacy /en shell: div.site-width announcement body + modal fragments.',
    'FAQ Page':
        'Legacy /en shell: div.site-width FAQ list.',
    'Info Index':
        'Legacy /en shell: div.site-width info link list.',
    'HTTP Error Page':
        'Legacy /en shell: simple error message wrapper (replace with native EDS error handling).',
    'Member Dining-Selection Detail (R####)':
        'Auth-gated/JS-rendered: no stable server DOM (redirects to JS search index; '
        'input-heavy member UI). Data-driven, not a static page DOM.',
    'Member Dining-Selection Index/Search':
        'Auth-gated member search index (JS-driven list/filter).',
    'SSI Include Fragment':
        'BARE PARTIAL — empty <head>, no <title>; only snippet markup (e.g. div.footerBlock / '
        'div.footerQABlock). This IS the header/footer/side-nav content, not a page body.',
    'LP Partial Fragment':
        'BARE PARTIAL — LP _offer/_header/_footer/parts snippet, assembled into a parent LP.',
    'Bumper/Redirect Stub': 'Interstitial bumper-link stub (no main content).',
    'Blank/Helper Stub': 'Blank iframe / javascript-off helper (no meaningful DOM).',
    'Tracking/System Stub': 'Tracking/log/verification endpoint (no page DOM).',
    'Modal/Popup Fragment': 'Modal/popup markup snippet loaded into a parent page.',
    'Auth Widget Fragment': 'Auth widget snippet (virtual numpad / button), not a page.',
    'Redirect / Router Stub': 'JS/meta redirect router stub (no content body).',
    'App-bridge Stub': 'Native-app bridge endpoint (no content body).',
    'System Error Stub': 'Point-mall system-failure stub (no content body).',
    'E-statement Email Template':
        'Email markup (noindex/nofollow): div.layout-canvas-a > banner + content + contentinfo. '
        'No site header/footer XF — email body, not a website page.',
}
DEFAULT_DOM = f'{AEM_ROOT} > standard AEM content blocks (confirm individually).'

# Styling helpers
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical='center', wrap_text=True)

MODE_FILL = {
    'Automated': 'C6EFCE',
    'Assisted': 'FFEB9C',
    'Manual': 'FFC7CE',
    'Manual / Replace': 'FFC7CE',
    'Manual / Auth-gated': 'F4CCCC',
    'Exclude (not a page)': 'D9D9D9',
    'Exclude / Re-point': 'D9D9D9',
    'Exclude / Separate track': 'D9D9D9',
}


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 28


wb = openpyxl.Workbook()

# --- Sheet 1: URL Mapping ---
ws = wb.active
ws.title = 'URL Mapping'
ws.append(['#', 'URL', 'Category', 'Migration Mode', 'Template Family',
           'Main Content DOM Pattern (between header & footer XF)'])
for i, r in enumerate(sorted(rows, key=lambda x: x['url']), start=1):
    fam = cat_to_family.get(r['category'], 'Other')
    dom = DOM_PATTERN.get(r['category'], DEFAULT_DOM)
    ws.append([i, r['url'], r['category'], r['mode'], fam, dom])
    mode_cell = ws.cell(row=i + 1, column=4)
    fill = MODE_FILL.get(r['mode'])
    if fill:
        mode_cell.fill = PatternFill('solid', fgColor=fill)
    for c in range(1, 7):
        ws.cell(row=i + 1, column=c).border = BORDER
        ws.cell(row=i + 1, column=c).alignment = WRAP
style_header(ws, 6)
ws.auto_filter.ref = f'A1:F{ws.max_row}'
widths = [6, 72, 32, 22, 34, 70]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 2: Category Summary ---
ws2 = wb.create_sheet('Category Summary')
ws2.append(['Category', 'Migration Mode', 'Template Family', 'URL Count',
            'Main Content DOM Pattern (between header & footer XF)'])
cat_counts = {}
cat_mode = {}
for r in rows:
    cat_counts[r['category']] = cat_counts.get(r['category'], 0) + 1
    cat_mode[r['category']] = r['mode']
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    ws2.append([cat, cat_mode[cat], cat_to_family.get(cat, 'Other'), n,
                DOM_PATTERN.get(cat, DEFAULT_DOM)])
    rr = ws2.max_row
    fill = MODE_FILL.get(cat_mode[cat])
    if fill:
        ws2.cell(row=rr, column=2).fill = PatternFill('solid', fgColor=fill)
    for c in range(1, 6):
        ws2.cell(row=rr, column=c).border = BORDER
        ws2.cell(row=rr, column=c).alignment = WRAP
ws2.append(['TOTAL (distinct URLs)', '', '', sum(cat_counts.values()), ''])
tot_row = ws2.max_row
for c in range(1, 6):
    ws2.cell(row=tot_row, column=c).font = Font(bold=True)
style_header(ws2, 5)
for i, w in enumerate([32, 22, 34, 12, 70], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 3: Mode Summary ---
ws3 = wb.create_sheet('Mode Summary')
ws3.append(['Migration Mode', 'URL Count', '% of analysed'])
mode_counts = {}
for r in rows:
    mode_counts[r['mode']] = mode_counts.get(r['mode'], 0) + 1
total = sum(mode_counts.values())
for m, n in sorted(mode_counts.items(), key=lambda x: -x[1]):
    ws3.append([m, n, f'{100*n/total:.1f}%'])
    rr = ws3.max_row
    fill = MODE_FILL.get(m)
    if fill:
        ws3.cell(row=rr, column=1).fill = PatternFill('solid', fgColor=fill)
    for c in range(1, 4):
        ws3.cell(row=rr, column=c).border = BORDER
ws3.append(['TOTAL', total, '100%'])
for c in range(1, 4):
    ws3.cell(row=ws3.max_row, column=c).font = Font(bold=True)
style_header(ws3, 3)
for i, w in enumerate([28, 12, 14], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Note row about R#### set
ws3.append([])
ws3.append(['Note: The 1,026 member dining-selection detail pages (…/search/R####.html) are now '
            'included as individual rows under the single "Member Dining-Selection Detail (R####)" '
            'category (Manual / Auth-gated). They share one auth-gated, JS-rendered template, so '
            'although counted individually they represent one build, not 1,026 distinct templates.'])
ws3.merge_cells(start_row=ws3.max_row, start_column=1, end_row=ws3.max_row, end_column=3)
ws3.cell(row=ws3.max_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
ws3.row_dimensions[ws3.max_row].height = 60

wb.save(OUT)
print('Saved:', OUT)
print('URL Mapping rows:', len(rows), '| categories:', len(cat_counts), '| modes:', len(mode_counts))
