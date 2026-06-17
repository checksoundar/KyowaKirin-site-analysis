#!/usr/bin/env python3
"""Build the diners.co.jp URL categorization Excel: per-URL mapping with
Category, Migration Mode, Template Family and Main Content DOM Pattern columns,
plus Category and Mode summary sheets.

Same SMTC Adobe Experience Manager platform as sumitclub.jp; standard content
pages render their main content in <div class="root responsivegrid"> between the
header and footer Experience Fragments.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PER_URL = '/workspace/current/.migration/dinersJP/per-url.json'
OUT = '/workspace/current/DinersJP_URL_Category_Mapping.xlsx'

rows = json.load(open(PER_URL))

FAMILY = {
    'Standard AEM Content Shell (/ja/)': [
        'Magazine Article', 'Magazine Category Index', 'Magazine Index/Listing',
        'Press / News Detail', 'Press / News Index', 'Event Report Detail',
        'Event Detail', 'Event Index/Nav', 'Service / How-to / Category Page',
        'Section Landing Page', 'Campaign/Event Landing', 'Campaign/Event Index',
        'Benefit Detail/Listing', 'Benefit Index', 'Club Online Promo (unique)',
        'Legal / Policy Text', 'Contact / Support', 'FAQ Page', 'Sitemap',
        'About / History Page', 'Topic Page', 'Other Content Page', 'Section Index',
        'Homepage', 'Search Page',
    ],
    'AEM Card-Detail / Listing Shell': [
        'Card Product Detail', 'Card Lineup Listing/Index',
    ],
    'AEM Corporate (business) Shell': [
        'Corporate Top/Landing', 'Corporate Service Page', 'Corporate Privilege Page',
        'Corporate Card Detail', 'Corporate Card Listing', 'Corporate Club Online How-to',
        'Merchant Top/Landing', 'Merchant Service Page',
    ],
    'SEO Article Microsite (entry_form/corporate)': [
        'Corporate Card SEO/Oyakudachi Article',
    ],
    'Application LP Microsite (entry_form)': [
        'Card Application Landing Page (LP)',
    ],
    'Standalone Microsites (non-AEM)': [
        'Identity Verification Microsite', 'Cancellation/Termination Flow',
        'Ginza Restaurant Shop Page', 'Travel Guide Article', 'Lounge Display Screen',
        'Short-name Landing/Redirect',
    ],
    'Auth-gated Member Area (excluded)': [
        'Premium Member Page (auth-gated)', 'Premium Member Landing',
        'Sign-on / Login System',
    ],
    'System / Redirect (excluded)': [
        'Redirect / Router Stub', 'HTTP Error Page', 'External Booking Redirect',
        'Thank-you / Mail Stub',
    ],
    'Non-Page Assets (excluded)': [
        'LP Partial Fragment', 'SSI Include Fragment', 'Verification Partial Fragment',
        'Modal/Popup Fragment', 'Blank/Helper Stub', 'Bumper/Redirect Stub',
        'Test/Dev Stub', 'Random-hash Stub', 'Site-verification Stub',
    ],
}
cat_to_family = {c: f for f, cs in FAMILY.items() for c in cs}

AEM_ROOT = 'div.root.responsivegrid'
DOM_PATTERN = {
    'Homepage': f'{AEM_ROOT} > hero carousel (CCM020) + audience-segment nav grid + important-notice ribbon + dual news feeds + promo tiles. One-off layout.',
    'Magazine Article': f'{AEM_ROOT} > breadcrumb + article title + hero image + rich-text body + inline images + related-article list + category tags. Highly uniform.',
    'Magazine Category Index': f'{AEM_ROOT} > breadcrumb + h1 + filter + card grid of articles (load-more).',
    'Magazine Index/Listing': f'{AEM_ROOT} > h1 + category tiles + latest-article card grid.',
    'Press / News Detail': f'{AEM_ROOT} > breadcrumb + date + title + rich-text body + optional table/attachment.',
    'Press / News Index': f'{AEM_ROOT} > h1 + dated news list with load-more.',
    'Event Report Detail': f'{AEM_ROOT} > breadcrumb + title + event photos + rich-text report body.',
    'Event Detail': f'{AEM_ROOT} > breadcrumb + hero + event description + schedule/CTA.',
    'Event Index/Nav': f'{AEM_ROOT} > h1 + event card grid / nav banners.',
    'Service / How-to / Category Page': f'{AEM_ROOT} > breadcrumb + side sub-nav (CCM027) + h1 + image+text sections + step blocks + accordion + CTA (travel/gourmet/golf/usage/point/etc).',
    'Section Landing Page': f'{AEM_ROOT} > breadcrumb + h1 + category card grid + intro.',
    'Section Index': f'{AEM_ROOT} > h1 + section link list / card grid.',
    'Campaign/Event Landing': f'{AEM_ROOT} > hero + campaign detail blocks + CTA to apply/external.',
    'Campaign/Event Index': f'{AEM_ROOT} > h1 + campaign card grid.',
    'Benefit Detail/Listing': f'{AEM_ROOT} > breadcrumb + benefit gallery + benefit tables + store-info table + CTA (bespoke benefit template).',
    'Benefit Index': f'{AEM_ROOT} > h1 + benefit card grid + filter.',
    'Club Online Promo (unique)': f'{AEM_ROOT} > promo image+text blocks + CTA (Club Online context).',
    'Legal / Policy Text': f'{AEM_ROOT} > breadcrumb + h1 + numbered headings + paragraphs + tables (text-heavy).',
    'Contact / Support': f'{AEM_ROOT} > h1 + inquiry-type accordion + gateway buttons to external form / FAQ / phone (no native form).',
    'FAQ Page': f'{AEM_ROOT} > h1 + FAQ accordion / links.',
    'Sitemap': f'{AEM_ROOT} > h1 + hierarchical link lists.',
    'About / History Page': f'{AEM_ROOT} > breadcrumb + h1 + story/history content sections.',
    'Topic Page': f'{AEM_ROOT} > h1 + topic content sections.',
    'Other Content Page': f'{AEM_ROOT} > h1 + mixed rich-text/card sections (standard AEM shell).',
    'Search Page': f'{AEM_ROOT} > search box + results list (external search.diners.co.jp engine).',
    'Card Product Detail': f'{AEM_ROOT} > breadcrumb + card hero image + feature/benefit grid + fee/spec tables + comparison + apply CTA to external entry form. Richest template.',
    'Card Lineup Listing/Index': f'{AEM_ROOT} > breadcrumb + h1 + card comparison grid/tables + filters.',
    'Corporate Top/Landing': f'{AEM_ROOT} (business-site nav) > hero + audience segment cards + service grid.',
    'Corporate Service Page': f'{AEM_ROOT} (business-site nav) > breadcrumb + h1 + service sections + tables/CTA.',
    'Corporate Privilege Page': f'{AEM_ROOT} (business-site nav) > breadcrumb + h1 + privilege card grid + detail sections.',
    'Corporate Card Detail': f'{AEM_ROOT} (business-site nav) > breadcrumb + card hero + spec/fee tables + apply CTA.',
    'Corporate Card Listing': f'{AEM_ROOT} (business-site nav) > h1 + corporate card comparison grid.',
    'Corporate Club Online How-to': f'{AEM_ROOT} (business-site nav) > breadcrumb + h1 + numbered how-to screenshots/steps.',
    'Merchant Top/Landing': f'{AEM_ROOT} (merchant nav) > hero + merchant service cards.',
    'Merchant Service Page': f'{AEM_ROOT} (merchant nav) > breadcrumb + h1 + service/info sections + tables.',
    'Corporate Card SEO/Oyakudachi Article': 'entry_form/corporate SEO article microsite: bespoke article wrapper (not the AEM responsivegrid) > h1 + long-form rich text + related-article list + apply CTA.',
    'Card Application Landing Page (LP)': 'Standalone marketing LP microsite: bespoke divs (mainvisual + offer + contents + footer partials assembled), forms/QS widgets, tables. NOT the AEM shell.',
    'Identity Verification Microsite': 'Static no-js microsite: div.wrapper + form scaffolding (assembled from honninkakunin/include/* partials). Not AEM shell.',
    'Cancellation/Termination Flow': 'Minimal-chrome flow microsite: wrapper div + step content + CTA.',
    'Ginza Restaurant Shop Page': 'Ginza restaurant microsite: bespoke shop layout > shop photo + description + map/access + reservation CTA (separate template, not main AEM shell).',
    'Travel Guide Article': 'Travel-guides microsite: district/city guide layout > hero + POI sections + map (separate codebase, own SSI footer).',
    'Lounge Display Screen': 'Digital signage display page (lounge availability) — minimal chrome, auto-refresh status.',
    'Short-name Landing/Redirect': 'Short vanity URL (company/biz/BMW/jc) — landing or redirect to a campaign/section.',
    'Premium Member Page (auth-gated)': 'Behind member login (/premium/member): AEM content but not anonymously reachable; mirrors public service pages for premium members.',
    'Premium Member Landing': 'Premium member area entry/landing (auth-gated).',
    'Sign-on / Login System': 'Authentication entry point — external system.',
    'Redirect / Router Stub': 'rd/ , rp*_rd, to/redirect, premium*_rd router stub (JS/meta redirect) — re-pointed, not migrated.',
    'HTTP Error Page': 'Platform error page (4xx/5xx) — replace with native EDS error handling.',
    'External Booking Redirect': 'GDO golf booking login/redirect — external system.',
    'Thank-you / Mail Stub': 'Thank-you / email-link landing stub.',
    'LP Partial Fragment': 'BARE PARTIAL — LP _offer/_header/_footer/_mainvisual/_contents/qs_parts snippet assembled into a parent entry_form LP. Not a standalone page.',
    'SSI Include Fragment': 'BARE PARTIAL — server-side include (header/footer/log) snippet; empty <head>, markup only. This is the header/footer content, not a page.',
    'Verification Partial Fragment': 'BARE PARTIAL — honninkakunin include (body/header/footer/numpad/modal) snippet assembled into the verification microsite.',
    'Modal/Popup Fragment': 'Modal/popup error markup loaded into a parent LP.',
    'Blank/Helper Stub': 'Blank iframe helper page (no meaningful DOM).',
    'Bumper/Redirect Stub': 'Interstitial bumper-link stub.',
    'Test/Dev Stub': 'JS test / hash-file dev stub (not content).',
    'Random-hash Stub': 'Random-hash filename stub (verification/placeholder, no real content).',
    'Site-verification Stub': 'Google site-verification file.',
}
DEFAULT_DOM = f'{AEM_ROOT} > standard AEM content blocks (confirm individually).'

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical='center', wrap_text=True)
MODE_FILL = {
    'Automated': 'C6EFCE', 'Assisted': 'FFEB9C', 'Manual': 'FFC7CE',
    'Manual / Replace': 'FFC7CE', 'Manual / Auth-gated': 'F4CCCC',
    'Exclude (not a page)': 'D9D9D9', 'Exclude / Re-point': 'D9D9D9',
}


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 28


wb = openpyxl.Workbook()

# Sheet 1: URL Mapping
ws = wb.active
ws.title = 'URL Mapping'
ws.append(['#', 'URL', 'Category', 'Migration Mode', 'Template Family',
           'Main Content DOM Pattern (between header & footer XF)'])
for i, r in enumerate(sorted(rows, key=lambda x: x['url']), start=1):
    fam = cat_to_family.get(r['category'], 'Other')
    dom = DOM_PATTERN.get(r['category'], DEFAULT_DOM)
    ws.append([i, r['url'], r['category'], r['mode'], fam, dom])
    fill = MODE_FILL.get(r['mode'])
    if fill:
        ws.cell(row=i + 1, column=4).fill = PatternFill('solid', fgColor=fill)
    for c in range(1, 7):
        ws.cell(row=i + 1, column=c).border = BORDER
        ws.cell(row=i + 1, column=c).alignment = WRAP
style_header(ws, 6)
ws.auto_filter.ref = f'A1:F{ws.max_row}'
for i, w in enumerate([6, 72, 34, 22, 36, 72], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Sheet 2: Category Summary
ws2 = wb.create_sheet('Category Summary')
ws2.append(['Category', 'Migration Mode', 'Template Family', 'URL Count',
            'Main Content DOM Pattern (between header & footer XF)'])
cat_counts, cat_mode = {}, {}
for r in rows:
    cat_counts[r['category']] = cat_counts.get(r['category'], 0) + 1
    cat_mode[r['category']] = r['mode']
for cat, n in sorted(cat_counts.items(), key=lambda x: -x[1]):
    ws2.append([cat, cat_mode[cat], cat_to_family.get(cat, 'Other'), n,
                DOM_PATTERN.get(cat, DEFAULT_DOM)])
    rr = ws2.max_row
    fill = MODE_FILL.get(cat_mode[cat])
    if fill:
        ws2.cell(row=rr, column=2).fill = PatternFill('solid', fgColor=fill)
    for c in range(1, 6):
        ws2.cell(row=rr, column=c).border = BORDER
        ws2.cell(row=rr, column=c).alignment = WRAP
ws2.append(['TOTAL (distinct URLs)', '', '', sum(cat_counts.values()), ''])
for c in range(1, 6):
    ws2.cell(row=ws2.max_row, column=c).font = Font(bold=True)
style_header(ws2, 5)
for i, w in enumerate([34, 22, 36, 12, 72], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Sheet 3: Mode Summary
ws3 = wb.create_sheet('Mode Summary')
ws3.append(['Migration Mode', 'URL Count', '% of analysed'])
mode_counts = {}
for r in rows:
    mode_counts[r['mode']] = mode_counts.get(r['mode'], 0) + 1
total = sum(mode_counts.values())
for m, n in sorted(mode_counts.items(), key=lambda x: -x[1]):
    ws3.append([m, n, f'{100*n/total:.1f}%'])
    fill = MODE_FILL.get(m)
    if fill:
        ws3.cell(row=ws3.max_row, column=1).fill = PatternFill('solid', fgColor=fill)
    for c in range(1, 4):
        ws3.cell(row=ws3.max_row, column=c).border = BORDER
ws3.append(['TOTAL', total, '100%'])
for c in range(1, 4):
    ws3.cell(row=ws3.max_row, column=c).font = Font(bold=True)
style_header(ws3, 3)
for i, w in enumerate([30, 12, 14], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print('Saved:', OUT)
print('URL Mapping rows:', len(rows), '| categories:', len(cat_counts), '| modes:', len(mode_counts))
